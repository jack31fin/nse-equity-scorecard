# excel_dashboard.py
# Phase 6 — Builds the final formatted Excel dashboard from final_scorecard.csv

import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
PROCESSED_PATH = os.path.join(PROJECT_ROOT, "data", "processed")
OUTPUT_PATH = os.path.join(PROJECT_ROOT, "output")
os.makedirs(OUTPUT_PATH, exist_ok=True)

scorecard_path = os.path.join(PROCESSED_PATH, "final_scorecard.csv")
if not os.path.exists(scorecard_path):
    raise SystemExit("❌ final_scorecard.csv not found in data/processed/. Run scorecard.py (Phase 5) first.")

df = pd.read_csv(scorecard_path)

# ── Columns as actually produced by scorecard.py ──────────────────
COLUMNS_TO_SHOW = ["Stock", "Close", "FA_Score", "TA_Score", "Total_Score",
                   "Signal", "Conviction", "Reasoning"]
HEADER_LABELS = {
    "Stock": "Stock", "Close": "Current Price", "FA_Score": "FA Score",
    "TA_Score": "TA Score", "Total_Score": "Total Score",
    "Signal": "Final Signal", "Conviction": "Conviction", "Reasoning": "Reasoning",
}

missing_cols = [c for c in COLUMNS_TO_SHOW if c not in df.columns]
if missing_cols:
    raise SystemExit(
        f"❌ final_scorecard.csv is missing column(s): {missing_cols}\n"
        f"   Columns actually present: {df.columns.tolist()}\n"
        f"   Re-run scorecard.py, or check it matches the version that produced this file."
    )

display_df = df[COLUMNS_TO_SHOW].copy()

wb = Workbook()
ws = wb.active
ws.title = "Scorecard"

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SIGNAL_FILLS = {
    "BUY": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "HOLD": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "SELL": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"), bottom=Side(style="thin", color="D1D5DB"),
)
SIGNAL_COL_INDEX = COLUMNS_TO_SHOW.index("Signal") + 1
LEFT_ALIGN_COLS = {COLUMNS_TO_SHOW.index("Conviction") + 1, COLUMNS_TO_SHOW.index("Reasoning") + 1}

for col_idx, col_name in enumerate(display_df.columns, start=1):
    cell = ws.cell(row=1, column=col_idx, value=HEADER_LABELS.get(col_name, col_name))
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER

for row_idx, row in enumerate(display_df.itertuples(index=False), start=2):
    for col_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(
            horizontal="left" if col_idx in LEFT_ALIGN_COLS else "center", vertical="center"
        )
        if col_idx == SIGNAL_COL_INDEX:
            cell.fill = SIGNAL_FILLS.get(str(value), PatternFill())
            cell.font = Font(bold=True)

widths = [18, 13, 10, 10, 12, 13, 17, 48]
for col_idx, width in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 22

n_rows = len(display_df)
total_score_col = COLUMNS_TO_SHOW.index("Total_Score") + 1
stock_col = COLUMNS_TO_SHOW.index("Stock") + 1

chart = BarChart()
chart.title = "Total Score by Stock (out of 100)"
chart.y_axis.title = "Total Score"
chart.x_axis.title = "Stock"
chart.style = 10
chart.width = 24
chart.height = 11

data_ref = Reference(ws, min_col=total_score_col, min_row=1, max_row=n_rows + 1)
cats_ref = Reference(ws, min_col=stock_col, min_row=2, max_row=n_rows + 1)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
ws.add_chart(chart, "J2")

out_file = os.path.join(OUTPUT_PATH, "scorecard.xlsx")
wb.save(out_file)

print(f"✅ Excel dashboard saved to {out_file}")
print(f"   {n_rows} stocks, color-coded by signal, with embedded bar chart.")